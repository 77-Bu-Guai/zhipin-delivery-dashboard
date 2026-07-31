import json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, PieChart, Reference

with open('scripts/analysis_result.json', 'r', encoding='utf-8') as f:
    data = json.load(f)

wb = Workbook()

# ===== Color definitions =====
HEADER_FILL = PatternFill('solid', fgColor='1F4E79')
HEADER_FONT = Font(name='Arial', bold=True, color='FFFFFF', size=11)
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='1F4E79')
SUBTITLE_FONT = Font(name='Arial', bold=True, size=12, color='2E75B6')
BODY_FONT = Font(name='Arial', size=10)
RED_FONT = Font(name='Arial', size=10, color='CC0000', bold=True)
GREEN_FONT = Font(name='Arial', size=10, color='006600', bold=True)
YELLOW_FILL = PatternFill('solid', fgColor='FFF2CC')
LIGHT_RED_FILL = PatternFill('solid', fgColor='FCE4D6')
LIGHT_GREEN_FILL = PatternFill('solid', fgColor='E2EFDA')
LIGHT_BLUE_FILL = PatternFill('solid', fgColor='D6E4F0')
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)

def style_header(ws, row, cols, fill=HEADER_FILL, font=HEADER_FONT):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = font
        cell.fill = fill
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = THIN_BORDER

def style_data_row(ws, row, cols, alt=False):
    for c in range(1, cols+1):
        cell = ws.cell(row=row, column=c)
        cell.font = BODY_FONT
        cell.border = THIN_BORDER
        cell.alignment = Alignment(vertical='center', wrap_text=True)
        if alt:
            cell.fill = PatternFill('solid', fgColor='F2F7FB')

# ==========================================
# Sheet 1: 总览仪表盘
# ==========================================
ws1 = wb.active
ws1.title = '总览仪表盘'
ws1.sheet_properties.tabColor = '1F4E79'

s = data['summary']
sd = data['scoreDistribution']

# Title
ws1.merge_cells('A1:F1')
ws1['A1'] = 'BOSS直聘投递数据分析报告 — 提示词优化依据'
ws1['A1'].font = TITLE_FONT

ws1.merge_cells('A2:F2')
ws1['A2'] = f'数据来源: extension-data.json | 导出时间: 2026-07-28 | 分析时间: 2026-07-29'
ws1['A2'].font = Font(name='Arial', size=9, color='666666')

# Key metrics
row = 4
metrics = [
    ('A', '总投递数', s['totalPipeline'], '条', LIGHT_BLUE_FILL),
    ('C', '投递成功', s['successCount'], '条', LIGHT_GREEN_FILL),
    ('E', '成功率', f"{s['successRate']}%", '', LIGHT_GREEN_FILL),
    ('A', 'AI评分日志', s['aiScoringLogs'], '条', LIGHT_BLUE_FILL),
    ('C', '被AI过滤', s['warningCount'], '条', LIGHT_RED_FILL),
    ('E', '过滤率', f"{s['filterRate']}%", '', LIGHT_RED_FILL),
]
for col, label, val, unit, fill in metrics:
    cell = ws1[f'{col}{row}']
    cell.value = label
    cell.font = Font(name='Arial', size=10, color='666666')
    cell = ws1[f'{col}{row+1}']
    cell.value = f'{val}{unit}'
    cell.font = Font(name='Arial', bold=True, size=18)
    cell.fill = fill

row = 7
ws1.merge_cells('A7:F7')
ws1['A7'] = '评分统计'
ws1['A7'].font = SUBTITLE_FONT

# Score stats table
row = 8
score_headers = ['指标', '成功条目', '过滤条目']
for i, h in enumerate(score_headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header(ws1, row, 3)

score_rows = [
    ('平均分', f"{s['successAvgScore']}", f"{s['warningAvgScore']}"),
    ('最低分', f"{s['successScoreMin']}", f"{s['warningScoreMin']}"),
    ('最高分', f"{s['successScoreMax']}", f"{s['warningScoreMax']}"),
]
for i, (label, sc, wc) in enumerate(score_rows):
    r = row + 1 + i
    ws1.cell(row=r, column=1, value=label)
    ws1.cell(row=r, column=2, value=sc)
    ws1.cell(row=r, column=3, value=wc)
    style_data_row(ws1, r, 3, i%2==1)
    ws1.cell(row=r, column=2).font = GREEN_FONT
    ws1.cell(row=r, column=3).font = RED_FONT

# Score distribution table
row = 13
ws1.merge_cells('A13:F13')
ws1['A13'] = '分数分布'
ws1['A13'].font = SUBTITLE_FONT

row = 14
dist_headers = ['分数区间', '成功数', '成功占比', '', '分数区间', '过滤数', '过滤占比']
for i, h in enumerate(dist_headers, 1):
    if h:
        ws1.cell(row=row, column=i, value=h)
style_header(ws1, row, 7)
ws1.cell(row=row, column=4).fill = PatternFill()  # empty middle column

succ_total = sum(sd['success'].values())
warn_total = sum(sd['warning'].values())
succ_keys = list(sd['success'].keys())
warn_keys = list(sd['warning'].keys())

for i in range(max(len(succ_keys), len(warn_keys))):
    r = row + 1 + i
    if i < len(succ_keys):
        ws1.cell(row=r, column=1, value=succ_keys[i])
        v = sd['success'][succ_keys[i]]
        ws1.cell(row=r, column=2, value=v)
        ws1.cell(row=r, column=3, value=f"{v/succ_total*100:.1f}%")
    if i < len(warn_keys):
        ws1.cell(row=r, column=5, value=warn_keys[i])
        v = sd['warning'][warn_keys[i]]
        ws1.cell(row=r, column=6, value=v)
        ws1.cell(row=r, column=7, value=f"{v/warn_total*100:.1f}%")
    style_data_row(ws1, r, 7, i%2==1)

# Pipeline flow summary
row = 22
ws1.merge_cells('A22:F22')
ws1['A22'] = '流水线效率'
ws1['A22'].font = SUBTITLE_FONT

row = 23
flow_headers = ['状态', '数量', '占比', '说明']
for i, h in enumerate(flow_headers, 1):
    ws1.cell(row=row, column=i, value=h)
style_header(ws1, row, 4)

flows = [
    ('success', s['successCount'], '通过全部筛选，已投递', LIGHT_GREEN_FILL),
    ('warning', s['warningCount'], '被脚本自动过滤，未投递', LIGHT_RED_FILL),
    ('warn', s['warnCount'], '需人工审核（AI筛选/地址/活跃度等）', YELLOW_FILL),
    ('danger', s['dangerCount'], '异常错误（超时/限制）', LIGHT_RED_FILL),
]
for i, (status, cnt, desc, fill) in enumerate(flows):
    r = row + 1 + i
    ws1.cell(row=r, column=1, value=status)
    ws1.cell(row=r, column=2, value=cnt)
    ws1.cell(row=r, column=3, value=f"{cnt/s['totalPipeline']*100:.1f}%")
    ws1.cell(row=r, column=4, value=desc)
    style_data_row(ws1, r, 4, i%2==1)
    if status == 'success':
        ws1.cell(row=r, column=2).font = GREEN_FONT
    elif status in ('warning', 'danger'):
        ws1.cell(row=r, column=2).font = RED_FONT

# Column widths
for col in ['A','B','C','D','E','F']:
    ws1.column_dimensions[col].width = 20

# ==========================================
# Sheet 2: 负面因子分析
# ==========================================
ws2 = wb.create_sheet('负面因子TOP50')
ws2.sheet_properties.tabColor = 'C00000'

ws2.merge_cells('A1:G1')
ws2['A1'] = '负面扣分因子分析（按总扣分影响排序）'
ws2['A1'].font = TITLE_FONT

ws2.merge_cells('A2:G2')
ws2['A2'] = '分析这50个因子怎样命中岗位JD → 判断提示词扣分规则的合理性 → 识别哪些是真正的红线 vs AI过度敏感'
ws2['A2'].font = Font(name='Arial', size=9, color='666666', italic=True)

row = 4
neg_headers = ['排名', '扣分原因', '出现次数', '总扣分', '平均扣分', '最低扣分', '最高扣分']
for i, h in enumerate(neg_headers, 1):
    ws2.cell(row=row, column=i, value=h)
style_header(ws2, row, 7)

neg_data = data['topNegativeReasons'][:50]
for idx, item in enumerate(neg_data):
    r = row + 1 + idx
    ws2.cell(row=r, column=1, value=idx+1)
    ws2.cell(row=r, column=2, value=item['reason'])
    ws2.cell(row=r, column=3, value=item['count'])
    ws2.cell(row=r, column=4, value=item['totalScore'])
    ws2.cell(row=r, column=5, value=item['avgScore'])
    ws2.cell(row=r, column=6, value=item['minScore'])
    ws2.cell(row=r, column=7, value=item['maxScore'])
    style_data_row(ws2, r, 7, idx%2==1)
    
    # Color-code severity
    if item['avgScore'] >= 800:
        ws2.cell(row=r, column=2).font = RED_FONT
        ws2.cell(row=r, column=4).font = RED_FONT
    elif item['avgScore'] >= 300:
        ws2.cell(row=r, column=2).font = Font(name='Arial', size=10, color='FF6600')
    # Highlight consistency issues
    if item['minScore'] != item['maxScore'] and item['count'] > 5:
        ws2.cell(row=r, column=6).fill = YELLOW_FILL
        ws2.cell(row=r, column=7).fill = YELLOW_FILL

# Add analysis notes
r = row + 51
ws2.merge_cells(f'A{r}:G{r}')
ws2[f'A{r}'] = '【关键发现】'
ws2[f'A{r}'].font = Font(name='Arial', bold=True, size=11, color='C00000')

notes = [
    '1. "大小周/单休" 是最大杀手（216,500总扣分），AI对此识别准确，规则合理',
    '2. "CET-4/英语要求" 类扣分严重（合计约15万总扣分），但规则设定为-1000是否过重？用户有基础英语能力',
    '3. "不接受居家办公" 出现256次但每次仅扣10分，属于轻微提醒项，非筛选关键',
    '4. "跨境电商" 类出现约200次但每次仅扣50分，为中度警示，非红线',
    '5. 同一概念有大量同义变体（大小周=大小周制=大小周工作=周末大小周=公司是大小周...），AI未能归并',
    '6. "薪资面议"出现168次（轻微提醒），但"经验不限"也出现在负面列表中（评分混乱）',
    '7. "福利：undefined" 出现24次 → 模板变量泄漏bug，需修复',
]
for i, note in enumerate(notes):
    r2 = r + 1 + i
    ws2.merge_cells(f'A{r2}:G{r2}')
    ws2[f'A{r2}'] = note
    ws2[f'A{r2}'].font = BODY_FONT

for col, w in zip(['A','B','C','D','E','F','G'], [6, 50, 10, 10, 10, 10, 10]):
    ws2.column_dimensions[col].width = w

# ==========================================
# Sheet 3: 正面因子分析
# ==========================================
ws3 = wb.create_sheet('正面因子TOP50')
ws3.sheet_properties.tabColor = '006600'

ws3.merge_cells('A1:G1')
ws3['A1'] = '正面加分因子分析（按总加分影响排序）'
ws3['A1'].font = TITLE_FONT

row = 3
pos_headers = ['排名', '加分原因', '出现次数', '总加分', '平均加分', '最低加分', '最高加分']
for i, h in enumerate(pos_headers, 1):
    ws3.cell(row=row, column=i, value=h)
style_header(ws3, row, 7)

pos_data = data['topPositiveReasons'][:50]
for idx, item in enumerate(pos_data):
    r = row + 1 + idx
    ws3.cell(row=r, column=1, value=idx+1)
    ws3.cell(row=r, column=2, value=item['reason'])
    ws3.cell(row=r, column=3, value=item['count'])
    ws3.cell(row=r, column=4, value=item['totalScore'])
    ws3.cell(row=r, column=5, value=item['avgScore'])
    ws3.cell(row=r, column=6, value=item['minScore'])
    ws3.cell(row=r, column=7, value=item['maxScore'])
    style_data_row(ws3, r, 7, idx%2==1)
    if 'AI' in item['reason'] or '大模型' in item['reason'] or 'Agent' in item['reason']:
        ws3.cell(row=r, column=2).font = GREEN_FONT

r = row + 52
ws3.merge_cells(f'A{r}:G{r}')
ws3[f'A{r}'] = '【关键发现】'
ws3[f'A{r}'].font = Font(name='Arial', bold=True, size=11, color='006600')

pnotes = [
    '1. "深圳" 相关加分最高频（682+94+85+89+67=1017次），这是用户的核心地理优势',
    '2. "五险一金" 出现491次但每次仅+8-10分，作为基本福利加分偏低',
    '3. "经验不限" 出现在正面430次，但也出现在负面，提示词对"经验"的处理存在二义性',
    '4. AI/大模型相关岗位加分高（+50分/次），但出现频率低（仅49次），建议扩大AI相关关键词覆盖',
    '5. 项目管理类（项目助理、项目管理能力、周报月报复盘等）加分稳定，与用户画像高度匹配',
    '6. 正面因子平均分普遍偏低（大多8-10分），而负面因子动辄-1000，悬殊过大可能导致少量大扣分就压制大量小加分',
]
for i, note in enumerate(pnotes):
    r2 = r + 1 + i
    ws3.merge_cells(f'A{r2}:G{r2}')
    ws3[f'A{r2}'] = note
    ws3[f'A{r2}'].font = BODY_FONT

for col, w in zip(['A','B','C','D','E','F','G'], [6, 50, 10, 10, 10, 10, 10]):
    ws3.column_dimensions[col].width = w

# ==========================================
# Sheet 4: 评分一致性问题
# ==========================================
ws4 = wb.create_sheet('评分一致性分析')
ws4.sheet_properties.tabColor = 'FF6600'

ws4.merge_cells('A1:G1')
ws4['A1'] = '评分一致性分析 — 同一概念评分差异过大的因子'
ws4['A1'].font = TITLE_FONT

ws4.merge_cells('A2:G2')
ws4['A2'] = 'min≠max且count>5的因子，说明AI对同一概念理解不一致 → 提示词需要更明确的规则约束'
ws4['A2'].font = Font(name='Arial', size=9, color='666666', italic=True)

row = 4
cons_headers = ['扣分原因', '出现次数', '最低扣分', '最高扣分', '差值', '平均扣分', '严重程度']
for i, h in enumerate(cons_headers, 1):
    ws4.cell(row=row, column=i, value=h)
style_header(ws4, row, 7)

for idx, item in enumerate(data['consistencyIssues']):
    r = row + 1 + idx
    diff = item['max'] - item['min']
    severity = '严重' if diff >= 2000 else ('较高' if diff >= 500 else ('中等' if diff >= 100 else '轻微'))
    
    ws4.cell(row=r, column=1, value=item['reason'])
    ws4.cell(row=r, column=2, value=item['count'])
    ws4.cell(row=r, column=3, value=item['min'])
    ws4.cell(row=r, column=4, value=item['max'])
    ws4.cell(row=r, column=5, value=diff)
    ws4.cell(row=r, column=6, value=round(item['avg']))
    ws4.cell(row=r, column=7, value=severity)
    style_data_row(ws4, r, 7, idx%2==1)
    
    if severity == '严重':
        ws4.cell(row=r, column=7).font = RED_FONT
    elif severity == '较高':
        ws4.cell(row=r, column=7).font = Font(name='Arial', size=10, color='FF6600')

# Add chart for consistency
chart = BarChart()
chart.type = 'col'
chart.title = '评分差异范围（min-max）'
chart.y_axis.title = '扣分差异'
chart.x_axis.title = '因子'
chart.style = 10
chart.width = 25
chart.height = 12

data_ref = Reference(ws4, min_col=5, min_row=row, max_row=row+len(data['consistencyIssues']))
cats_ref = Reference(ws4, min_col=1, min_row=row+1, max_row=row+len(data['consistencyIssues']))
chart.add_data(data_ref, titles_from_data=True)
chart.set_categories(cats_ref)
chart.series[0].graphicalProperties.solidFill = "FF6600"

ws4.add_chart(chart, f'A{row+len(data["consistencyIssues"])+3}')

for col, w in zip(['A','B','C','D','E','F','G'], [45, 10, 10, 10, 8, 10, 8]):
    ws4.column_dimensions[col].width = w

# ==========================================
# Sheet 5: 边际成功分析
# ==========================================
ws5 = wb.create_sheet('边际成功案例')
ws5.sheet_properties.tabColor = 'FFC000'

ws5.merge_cells('A1:E1')
ws5['A1'] = '低分通过案例 — 评分0-55分但仍被投递的职位'
ws5['A1'].font = TITLE_FONT

ws5.merge_cells('A2:E2')
ws5['A2'] = '这些案例得分很低但通过了筛选 → 可能存在评分阈值设置问题，或这些岗位实际匹配度存疑'
ws5['A2'].font = Font(name='Arial', size=9, color='666666', italic=True)

row = 4
b_headers = ['排名', '岗位名称', '公司', '得分', '风险等级']
for i, h in enumerate(b_headers, 1):
    ws5.cell(row=row, column=i, value=h)
style_header(ws5, row, 5)

for idx, item in enumerate(data['borderlineSuccess']):
    r = row + 1 + idx
    risk = '高风险' if item['score'] <= 35 else ('中风险' if item['score'] <= 45 else '低风险')
    ws5.cell(row=r, column=1, value=idx+1)
    ws5.cell(row=r, column=2, value=item['job'])
    ws5.cell(row=r, column=3, value=item['company'])
    ws5.cell(row=r, column=4, value=item['score'])
    ws5.cell(row=r, column=5, value=risk)
    style_data_row(ws5, r, 5, idx%2==1)
    if risk == '高风险':
        ws5.cell(row=r, column=5).font = RED_FONT
    elif risk == '中风险':
        ws5.cell(row=r, column=5).font = Font(name='Arial', size=10, color='FF6600')

for col, w in zip(['A','B','C','D','E'], [6, 35, 25, 8, 10]):
    ws5.column_dimensions[col].width = w

# ==========================================
# Sheet 6: 问题诊断 & 优化建议
# ==========================================
ws6 = wb.create_sheet('问题诊断与优化建议')
ws6.sheet_properties.tabColor = '7030A0'

ws6.merge_cells('A1:F1')
ws6['A1'] = '提示词问题诊断 & 优化建议'
ws6['A1'].font = TITLE_FONT

row = 3
diag_headers = ['编号', '问题类别', '严重程度', '问题描述', '数据依据', '优化建议']
for i, h in enumerate(diag_headers, 1):
    ws6.cell(row=row, column=i, value=h)
style_header(ws6, row, 6)

issues = [
    ('P1', '模板变量泄漏', '严重', 
     '"福利"等变量显示为undefined，76条日志中存在此Bug',
     '76条含有"undefined"的评分，影响评分准确性',
     '检查前端模板渲染逻辑，变量为空时应跳过而非输出"undefined"'),
    ('P2', '同义项未归并', '严重',
     '"大小周"有至少15种变体表述（大小周制/大小周休息/大小周工作制/公司是大小周/周末大小周...），各变体分数也不一致（50~5000）',
     '15+种"大小周"变体，分数范围50-5000',
     '在提示词中增加同义词表/归并规则，或增加预处理步骤统一关键词'),
    ('P3', '英语能力扣分过重', '较高',
     'CET-4硬性要求扣-1000，英语工作语言扣-100~-1000。用户可阅读技术文档但无证书。230+条目因此被严重扣分',
     'CET/英语类扣分合计约15万分，影响230+条目',
     '细分英语要求等级（硬性CET-6 vs 基础读写 vs 全英文环境），基础英语能力仅扣-50~-100'),
    ('P4', '评分正负悬殊', '较高',
     '正面因子大多仅+8~+10分，负面因子动辄-1000。一个小负面就可抵消100个正面加分',
     '正面avg=10，负面avg重达-1000，比例1:100',
     '重新平衡评分权重。核心红线保持-1000，中度警示-50~-100，正面核心匹配提升至+100~+200'),
    ('P5', '分类错位', '较高',
     '710条记录中出现可疑的正面分类（销售/英语四级/CET等被归类为正面加分）',
     '710条可疑正面分类',
     '提示词中对"销售"类、"英语"类的判定规则有歧义，需明确：销售助理≠纯销售岗位，CET-4永远不应出现在正面'),
    ('P6', '低分通过', '较高',
     '66个职位得分0-55仍然投递成功（如招聘专员30分、客服专员30分）',
     '66个低分成功案例，与过滤阈值矛盾',
     '检查过滤阈值逻辑，成功/过滤的分界线是否预期。建议success阈值设在≥50'),
    ('P7', '福利项权重偏低', '中等',
     '"五险一金"出现491次但每次仅+8~+10分，"双休"出现412次仅+18~+19分',
     '核心福利加分仅占总分的个位数%',
     '大幅提升核心硬福利权重：双休+50，五险一金+30，年终奖+20'),
    ('P8', 'AI关键词覆盖不足', '中等',
     '"岗位涉及AI/智能体/大模型/Agent/RAG且非纯开发岗"仅命中49次，但AI相关描述应更多',
     '仅49次命中AI加分，明显偏少',
     '扩展AI关键词：增加LLM/机器学习/深度学习/NLP/计算机视觉/多模态/知识图谱等'),
    ('P9', '"抗压"过度解读', '中等',
     '"抗压能力强"68次/"抗压能力"48次被判为"狼性文化/抗压/能吃苦"扣50分，但很多JD的"抗压"只是泛泛之词',
     '116+次"抗压"触发扣分',
     '"抗压能力"不应自动=狼性文化。仅在明确出现"狼性/996/加班文化"等连用时才扣分'),
    ('P10', 'SD-Issue: 薪资范围匹配', '轻微',
     '"薪资面议"168次扣-10，但实际薪资在JD中可能已给出',
     '168次轻微扣分，价值不大',
     '"薪资面议"改为不加分不扣分，或仅在真的完全无薪资信息时扣分'),
]

for idx, (pid, cat, sev, desc, data_ref, suggestion) in enumerate(issues):
    r = row + 1 + idx
    ws6.cell(row=r, column=1, value=pid)
    ws6.cell(row=r, column=2, value=cat)
    ws6.cell(row=r, column=3, value=sev)
    ws6.cell(row=r, column=4, value=desc)
    ws6.cell(row=r, column=5, value=data_ref)
    ws6.cell(row=r, column=6, value=suggestion)
    style_data_row(ws6, r, 6, idx%2==1)
    ws6.row_dimensions[r].height = 50
    
    if sev == '严重':
        ws6.cell(row=r, column=3).font = RED_FONT
    elif sev == '较高':
        ws6.cell(row=r, column=3).font = Font(name='Arial', size=10, color='FF6600')

# Optimization priority summary
r = row + len(issues) + 2
ws6.merge_cells(f'A{r}:F{r}')
ws6[f'A{r}'] = '优化优先级排序（建议执行顺序）'
ws6[f'A{r}'].font = SUBTITLE_FONT

priorities = [
    '优先级1 → P1(修复undefined bug) + P3(调整英语扣分) + P7(提升福利权重) — 立即可做，影响面最大',
    '优先级2 → P2(归并同义词) + P4(平衡正负权重) — 需要重新设计评分体系，工作量中等',
    '优先级3 → P5(修复分类错位) + P6(调整过滤阈值) + P8(扩展AI关键词) — 精细调优',
    '优先级4 → P9(抗压解读) + P10(薪资面议) — 微调，影响面较小',
]
for i, p in enumerate(priorities):
    r2 = r + 1 + i
    ws6.merge_cells(f'A{r2}:F{r2}')
    ws6[f'A{r2}'] = p
    ws6[f'A{r2}'].font = BODY_FONT

for col, w in zip(['A','B','C','D','E','F'], [6, 18, 8, 45, 30, 45]):
    ws6.column_dimensions[col].width = w

# ==========================================
# Save Excel
# ==========================================
output_path = 'scripts/BOSS投递数据_提示词优化分析报告.xlsx'
wb.save(output_path)
print(f'Excel report saved: {output_path}')
